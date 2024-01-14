import smtplib, ssl

from email.message import EmailMessage

from openpyxl import load_workbook

### CONF VARIABLES ###

sender = "example@domain.com"

mdp = "very_comlicated"

host = "smtp_server_goes_here"

notes_file = "grades.xlsx" # ficier contenant les notes au format xlsx

note_ajoutee = "Exam1" # last added exam 

f_message = "message_grades.txt"

nb_exams = 2

##############

msgf = EmailMessage()

msgf["From"] = sender

port = 587 

#ctx = ssl._create_unverified_context()
#
#clt = smtplib.SMTP("{}:{}".format(host, port))
#
#clt.starttls(context=ctx)
#
#clt.login(sender, mdp)

notes_l = []

notesf_l = []

to_l = []

with open(f_message, "r") as file:

    msg = file.read()

db_notes = load_workbook(notes_file)

fl = db_notes.active

sujet = fl.cell(5, 1).value 

msgf["Subject"] = sujet 

msg = msg.replace("C", sujet)

cnt = 9 

flag_l = []

names_exam_l = []

[ flag_l.append(fl.cell(cnt, 3+i).value) for i in range(nb_exams) ]

[ names_exam_l.append(fl.cell(8, 3+i).value) for i in range(nb_exams) ]

idx_notes = names_exam_l.index(note_ajoutee) + 3

if flag_l.count("None") == len(flag_l):

    msg = msg.replace("F", "")

else:

    msg = msg.replace("F", ", for now,")

while fl.cell(cnt, 1).value != None:

    notes_l.append(fl.cell(cnt, idx_notes).value)

    to_l.append(fl.cell(cnt, nb_exams + 4).value)

    notesf_l.append(fl.cell(cnt, nb_exams + 3).value)

    cnt += 1

db_notes.close()

msg = msg.replace("X", sujet)

for i in range(len(to_l)):

    etudiant = to_l[i]

    msg_et = msg.replace("Y", str(notes_l[i]))

    msgf.set_content(msg_et.replace("A", notesf_l[i]))

    clt.sendmail(sender, etuduiant, msgf.as_string())

    print("Mail sent for: {}".format(etudiant))

clt.quit()

print("")

print("done")




